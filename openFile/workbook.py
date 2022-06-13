from openpyxl import load_workbook
import json


def get_json(path):
    book = load_workbook(path)
    sheet = book.active
    rows = sheet.rows
    sheets = book.sheetnames[0]

    headers = [cell.value for cell in next(rows)]

    all_rows = []

    persons = {
        sheets: all_rows
    }

    for row in rows:
        date = {}
        for title, cell in zip(headers, row):
            date[title] = cell.value
        all_rows.append(date)

    with open('persons.json', 'w') as file:
        json.dump(persons, file, indent=2, ensure_ascii=False)



