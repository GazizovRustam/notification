from openpyxl import load_workbook
import json


# def get_json(path):
book = load_workbook('Meta.xlsx')
sheet = book.active
rows = sheet.rows
sheets = book.sheetnames[0]

headers = [cell.value for cell in next(rows)]

all_rows = []

persons = {
    sheets: all_rows
}

for row in rows:
    date = {}
    for title, cell in zip(headers, row):
        date[title] = cell.value
    all_rows.append(date)


with open(sheets, 'w', encoding='utf8') as file:
    json.dump(persons, file, indent=2, ensure_ascii=False)

    # return persons

# get_json(r'C:\Users\GRX\Desktop\Meta.xlsx')