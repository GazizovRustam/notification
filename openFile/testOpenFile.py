import openpyxl as opx

book = opx.open('Meta.xlsx', read_only=True)
sheet = book.active
count_c = [0] * 10
count_r = [0] * sheet.max_row
#
def get_sheet():

    top = 0
    for column in range(0, sheet.max_column):
        date_column = sheet[1][column].value
        for x in range(0, 1):
            count_c[top] = date_column
            top += 1
    # print(count_c)
    #column = 0




top = 0
for column in range(0, sheet.max_column):
    for row in range(1, sheet.max_row + 1):
        date = sheet[row][column].value
        count_c[top] = date
        top += 1
        print(count_c)







    # for n in range(1, 5):
    #     for x in range(0, 1):
    #         for v in range(x, 1):
    #             person = {
    #                 count_r[x]: count_r[n],
    #                 count_r[v]: count_r[n]
    #             }
    #             print(person)


get_sheet()
# def main():







# column =0
# while column != sheet.max_column:
#     date_column = sheet[1][column].value
#     count_c[column] = date_column
#     column += 1
#
# print(count_c)
    #
    # while row != sheet.max_row:
    #     date_row = sheet[row][0].value
    #     count_r[top] = date_row
    #     top += 1
    #     row += 1
    #     print(count_r)


    # person = {
    #     count[0]: count[1],
    #     count[2]: count[3]
    # }
    # print(person)

        # count = [0]
        # count.append(date)
        # print(count)
    #num = [date]
    #print(num)



